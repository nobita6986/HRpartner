# -*- coding: utf-8 -*-
"""Count by hand: tổng '20h~22h' và 'Số giờ đêm' cho A602010881"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.load_workbook(r'C:\CodeApp\HrP\appBCC\docs\Actro\BCCActroT7.xlsx', data_only=True)
ws = wb.active

# Find Ma A Sình (A602010881)
target_code = 'A602010881'
emp_row = None
for row in range(3, ws.max_row + 1):
    c = ws.cell(row=row, column=2).value
    if c and str(c).strip() == target_code:
        emp_row = row
        break

print(f"Found {target_code} at row {emp_row}")
print(f"Name: {ws.cell(row=emp_row, column=3).value}")

# Find day columns from header row 8
day_cols = {}
for col in range(6, 290):
    val = ws.cell(row=8, column=col).value
    if val is not None and str(val).strip().isdigit():
        d = int(str(val).strip())
        if 1 <= d <= 31:
            day_cols[d] = col

# Determine period: ngày 26-31 thuộc tháng 6, ngày 1-25 thuộc tháng 7
def get_date_for_day(day_num):
    if day_num >= 26:
        return date(2026, 6, day_num)
    else:
        return date(2026, 7, day_num)

# Sub-header layout for each day
def get_sub_headers(start_col):
    headers = []
    for offset in range(10):
        h = ws.cell(row=9, column=start_col+offset).value
        headers.append(h)
    return headers

# Now manually count per day
print(f"\n{'='*100}")
print(f"{'Day':<5} | {'Date':<12} | {'Day-type':<10} | {'In':<7} | {'Out':<7} | {'20h22':<6} | {'Đêm130':<6} | {'200%đêm':<8} | {'210CN':<5}")
print("="*100)

total_20h22 = 0  # KE
total_dem_130 = 0  # KI
total_200_dem = 0  # KH

day_records = []

for day_num in sorted(day_cols.keys()):
    start = day_cols[day_num]
    sub_headers = get_sub_headers(start)
    
    in_t = ws.cell(row=emp_row, column=start).value
    out_t = ws.cell(row=emp_row, column=start+1).value
    hours = ws.cell(row=emp_row, column=start+2).value
    
    # Detect day type based on layout (day 28 has special layout: cols Z-AE)
    if start == 26:
        day_type = 'CN/Lễ'
    else:
        # Check if has 'In' sub-header (working day)
        if in_t and not isinstance(in_t, str):
            weekday = get_date_for_day(day_num).weekday()
            day_type = ['T2','T3','T4','T5','T6','T7','CN'][weekday]
        elif in_t == 'A' or in_t == 'D' or in_t == 'TV':
            day_type = in_t  # Absent/Day off/Compensatory
        else:
            day_type = '?'
    
    # Find the columns
    h_20_22 = 0
    h_dem_130 = 0
    h_200_dem = 0
    h_210_cn = 0
    
    for offset in range(10):
        if start + offset > 290:
            break
        sub = sub_headers[offset] if offset < len(sub_headers) else None
        val = ws.cell(row=emp_row, column=start+offset).value
        
        if val is None:
            continue
        if isinstance(val, str):
            continue
        
        if sub == '20h~22h':
            h_20_22 = val
            total_20h22 += val
        elif sub and 'Số giờ' in str(sub) and 'đêm' in str(sub):
            h_dem_130 = val
            total_dem_130 += val
        elif sub and '200%' in str(sub):
            h_200_dem = val
            total_200_dem += val
    
    in_str = str(in_t).split(' ')[-1] if in_t else ''
    out_str = str(out_t).split(' ')[-1] if out_t else ''
    
    day_date = get_date_for_day(day_num)
    
    print(f"{day_num:<5} | {day_date} | {day_type:<10} | {in_str:<7} | {out_str:<7} | {str(h_20_22):<6} | {str(h_dem_130):<6} | {str(h_200_dem):<8} | {str(h_210_cn):<5}")
    
    day_records.append({
        'day': day_num,
        'date': day_date,
        'day_type': day_type,
        'in': in_t,
        'out': out_t,
        'h_20_22': h_20_22,
        'h_dem_130': h_dem_130,
        'h_200_dem': h_200_dem,
    })

print("="*100)
print(f"\n=== TỔNG ===")
print(f"KE (20h~22h): {total_20h22}")
print(f"KI (Số giờ đêm 130%): {total_dem_130}")
print(f"KH (200% ca đêm): {total_200_dem}")

print(f"\n=== HR's values (for comparison) ===")
print(f"KE = 18 (HR)")
print(f"KI = 54 (HR)")
print(f"KH = 6 (HR)")

print(f"\n=== Analysis: where's the discrepancy? ===")
print(f"\nKE: BCC={total_20h22}, HR=18, diff={total_20h22 - 18}")
print(f"KI: BCC={total_dem_130}, HR=54, diff={total_dem_130 - 54}")

# Detail day-by-day showing both '20h~22h' and night hours
print(f"\n=== CHI TIẾT TỪNG NGÀY (chỉ các ngày có data) ===")
for rec in day_records:
    if rec['in'] and rec['day_type'] not in ['CN/Lễ']:
        # Tính số giờ đêm: giờ từ 22h-6h
        # In=20:00, Out=08:00 => đêm = 22h-24h(2h) + 0h-6h(6h) + 6h-8h(2h ngoài khung đêm)
        # Hoặc đơn giản: nếu ca đêm = 8h thì hết là đêm
        print(f"  Day {rec['day']} ({rec['date']} {rec['day_type']}): {rec['in']}→{rec['out']} | 20h22={rec['h_20_22']}h, đêm130={rec['h_dem_130']}h, 200%={rec['h_200_dem']}h")