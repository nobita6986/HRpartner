# -*- coding: utf-8 -*-
"""Debug employee A602010881"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.utils import get_column_letter

# HR data for A602010881
wb_hr = openpyxl.load_workbook(r'C:\CodeApp\HrP\appBCC\docs\Actro\LCNT7.xlsx', data_only=True)
ws_hr = wb_hr.active

for row_idx in range(12, ws_hr.max_row + 1):
    code = ws_hr.cell(row=row_idx, column=2).value
    if code and str(code).strip() == 'A602010881':
        print(f"=== HR Data for A602010881 (Row {row_idx}) ===")
        
        # Print all relevant fields
        lcb = ws_hr.cell(row=row_idx, column=304).value  # KR
        luong = ws_hr.cell(row=row_idx, column=305).value  # KS
        chuyen_can = ws_hr.cell(row=row_idx, column=306).value  # KT
        soi_kinh = ws_hr.cell(row=row_idx, column=307).value  # KU
        doi_song = ws_hr.cell(row=row_idx, column=308).value  # KV
        nha_o = ws_hr.cell(row=row_idx, column=309).value  # KW
        tham_nien = ws_hr.cell(row=row_idx, column=310).value  # KX
        thanh_toan = ws_hr.cell(row=row_idx, column=311).value  # KY
        tru_ung = ws_hr.cell(row=row_idx, column=312).value  # KZ
        thuc_nhan = ws_hr.cell(row=row_idx, column=313).value  # LA
        
        # OT data
        kc = ws_hr.cell(row=row_idx, column=289).value
        kd = ws_hr.cell(row=row_idx, column=290).value
        ke = ws_hr.cell(row=row_idx, column=291).value
        kf = ws_hr.cell(row=row_idx, column=292).value
        kh = ws_hr.cell(row=row_idx, column=294).value
        ki = ws_hr.cell(row=row_idx, column=295).value
        kj = ws_hr.cell(row=row_idx, column=296).value
        kk = ws_hr.cell(row=row_idx, column=297).value
        
        ngay_vao = ws_hr.cell(row=row_idx, column=4).value
        loai = ws_hr.cell(row=row_idx, column=299).value  # KM - Loại
        
        print(f"  Ngày vào: {ngay_vao}")
        print(f"  Loại: {loai}")
        print(f"  KX309: {kc}, KD290: {kd}, KE291: {ke}, KF292: {kf}, KH294: {kh}, KI295: {ki}, KJ296: {kj}, KK297: {kk}")
        print(f"  CÔNG HC (KL): {ws_hr.cell(row=row_idx, column=298).value}")
        print(f"  LCB/giờ: {lcb}")
        print(f"  LƯƠNG: {luong:,.0f}" if luong else "  LƯƠNG: None")
        print(f"  Chuyên cần: {chuyen_can}")
        print(f"  Soi kính: {soi_kinh}")
        print(f"  Đời sống: {doi_song}")
        print(f"  Nhà ở: {nha_o}")
        print(f"  Thâm niên: {tham_nien}")
        print(f"  Thanh toán: {thanh_toan:,.0f}" if thanh_toan else "  Thanh toán: None")
        print(f"  Trừ ứng: {tru_ung}")
        print(f"  Thực nhận: {thuc_nhan}")
        
        # Verify SUMPRODUCT formula
        lcb_hour = lcb if lcb else 6000000/26/8
        sumproduct = (kd or 0)*1 + (ke or 0)*1 + (kf or 0)*1.5 + (kh or 0)*2 + (ki or 0)*1.3 + (kj or 0)*2 + (kk or 0)*2.7
        calc_luong = lcb_hour * sumproduct
        print(f"\n  EXPECTED: SUMPRODUCT = {sumproduct}")
        print(f"  EXPECTED: LƯƠNG = {calc_luong:,.0f}")
        
        if luong:
            diff = abs(calc_luong - luong)
            print(f"  DIFF: {diff:,.0f}")
            print(f"  HR's actual LƯƠNG: {luong:,.0f}")
        
        # Check name
        name = ws_hr.cell(row=row_idx, column=3).value
        print(f"\n  Name: {name}")
        break