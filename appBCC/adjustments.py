"""Đọc file mẫu cộng/trừ do người dùng upload và áp vào dữ liệu lương.

Cấu trúc file:
- Sheet ``Danh sách khoản``: cột A = Mã NV, cột B = Họ tên.
- Từ cột C trở đi, các cột đi theo cặp: cột lẻ => cộng, cột chẵn => trừ.
- Tên cột được giữ nguyên do người dùng đổi, dùng làm tên khoản khi xuất.

Parser trả về ``dict[employee_code] -> list[Adjustment]``. Mỗi ``Adjustment`` có
``name``, ``amount`` và ``kind`` (PLUS/MINUS).
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl


ADJUSTMENT_KIND_PLUS = "PLUS"
ADJUSTMENT_KIND_MINUS = "MINUS"


@dataclass
class Adjustment:
    name: str
    amount: float
    kind: str  # PLUS hoặc MINUS


def _to_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_adjustment_file(file_path: str | Path) -> dict[str, list[Adjustment]]:
    """Đọc file cộng/trừ theo cặp cột (cột lẻ = cộng, cột chẵn = trừ)."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Không tìm thấy file: {file_path}")

    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet_name = next(
        (name for name in workbook.sheetnames if "khoản" in name.lower() or "khoan" in name.lower()),
        workbook.sheetnames[0],
    )
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    if len(header) < 3:
        raise ValueError("File mẫu cần ít nhất 2 cột Mã NV/Họ tên và 1 cặp cộng/trừ.")

    payload: dict[str, list[Adjustment]] = {}
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        employee_code = str(row[0]).strip()
        if not employee_code:
            continue

        adjustments: list[Adjustment] = []
        for offset, cell in enumerate(row[2:], start=0):
            amount = _to_float(cell)
            if amount == 0:
                continue
            column_index = 2 + offset  # index 0-based cho cột trong header
            column_name = header[column_index] if column_index < len(header) else f"Khoản {offset + 1}"
            if not column_name:
                column_name = f"Khoản {offset + 1}"
            if offset % 2 == 0:
                kind = ADJUSTMENT_KIND_PLUS
            else:
                kind = ADJUSTMENT_KIND_MINUS
            adjustments.append(Adjustment(name=column_name, amount=abs(amount), kind=kind))
        if adjustments:
            payload[employee_code] = adjustments

    workbook.close()
    return payload


def apply_adjustments_to_payroll(
    payroll_data: dict | None,
    adjustments: list[Adjustment],
) -> dict:
    """Gộp danh sách cộng/trừ vào kết quả payroll, giữ nguyên cấu trúc cũ."""
    if payroll_data is None:
        payroll_data = {
            "salaryItems": [],
            "allowances": [],
            "deductions": [],
            "summary": {
                "salaryNormal": 0,
                "thanhToan": 0,
                "totalDeduction": 0,
                "netIncome": 0,
            },
        }

    salary_items = list(payroll_data.get("salaryItems") or [])
    deductions = list(payroll_data.get("deductions") or [])

    total_plus = 0.0
    total_minus = 0.0
    for adjustment in adjustments:
        if adjustment.kind == ADJUSTMENT_KIND_PLUS:
            salary_items.append(
                {
                    "name": adjustment.name,
                    "qty": adjustment.amount,
                    "rate": 1,
                    "total": adjustment.amount,
                    "source": "adjustment_template",
                }
            )
            total_plus += adjustment.amount
        else:
            deductions.append(
                {
                    "name": adjustment.name,
                    "qty": None,
                    "rate": None,
                    "total": -adjustment.amount,
                    "source": "adjustment_template",
                }
            )
            total_minus += adjustment.amount

    summary = dict(payroll_data.get("summary") or {})
    summary["totalDeduction"] = float(summary.get("totalDeduction", 0) or 0) + total_minus
    summary["netIncome"] = float(summary.get("netIncome", 0) or 0) + total_plus - total_minus

    payroll_data["salaryItems"] = salary_items
    payroll_data["deductions"] = deductions
    payroll_data["summary"] = summary
    payroll_data["adjustments"] = [
        {"name": adjustment.name, "amount": adjustment.amount, "kind": adjustment.kind}
        for adjustment in adjustments
    ]
    return payroll_data
