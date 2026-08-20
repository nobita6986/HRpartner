# -*- coding: utf-8 -*-
"""Detailed BCC layout inspection"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.utils import get_column_letter

wb_bcc = openpyxl.load_workbook(r'C:\CodeApp\HrP\appBCC\docs\Actro\BCCActroT7.xlsx', data_only=True)
ws_bcc = wb_bcc.active

# Look at row 8 (header) to find where days 26, 27, 28 start
print("=== Row 8 (headers) ===")
for col in range(1, 50):
    val = ws_bcc.cell(row=8, column=col).value
    if val is not None:
        print(f"  Col {col} ({get_column_letter(col)}): {val}")

# Row 9 (sub-headers)
print("\n=== Row 9 (sub-headers) ===")
for col in range(1, 50):
    val = ws_bcc.cell(row=9, column=col).value
    if val is not None:
        print(f"  Col {col} ({get_column_letter(col)}): {val}")

# Row 13 (the row with day numbers) - find Ma Doan Chung first to confirm structure
# We need to know exact row of A602010881
for row in range(3, ws_bcc.max_row + 1):
    c = ws_bcc.cell(row=row, column=2).value
    if c and str(c).strip() == 'A602010881':
        print(f"\n=== A602010881 at row {row} ===")
        for col in range(1, ws_bcc.max_column + 1):
            val = ws_bcc.cell(row=row, column=col).value
            if val is not None:
                print(f"  Col {col} ({get_column_letter(col)}): {val}")
        break