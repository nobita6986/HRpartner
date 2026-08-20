# -*- coding: utf-8 -*-
"""Check the 'Overtime (2)' sheet in BCCActroT7_OK.xlsx - maybe employees are listed there"""
import sys
sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)

import openpyxl

INPUT = r"C:\CodeApp\HrP\appBCC\docs\Actro\round2\BCCActroT7_OK.xlsx"

wb = openpyxl.load_workbook(INPUT, data_only=True)
ws = wb["Overtime (2)"]

print(f"Sheet 'Overtime (2)': max_row={ws.max_row}, max_col={ws.max_column}")
print()

# First 15 rows
for r in range(1, 16):
    row_data = []
    for c in range(1, min(10, ws.max_column+1)):
        v = ws.cell(row=r, column=c).value
        row_data.append(str(v)[:30] if v is not None else "")
    print(f"  R{r}: {row_data}")

# Look for employees in Overtime (2) sheet
print("\n=== Employees in Overtime (2) sheet ===")
employees_ot = []
for r in range(2, ws.max_row + 1):
    code = ws.cell(row=r, column=2).value
    name = ws.cell(row=r, column=3).value
    if code and str(code).strip() and not str(code).strip().startswith("Mã"):
        employees_ot.append((r, str(code).strip(), str(name).strip() if name else ""))

print(f"Total: {len(employees_ot)}")
for r, code, name in employees_ot[:60]:
    print(f"  R{r}: {code} - {name}")

# Compare with output
OUTPUT = r"C:\CodeApp\HrP\appBCC\docs\Actro\round2\LCNT_7_2026_BCC.xlsx"
wb_out = openpyxl.load_workbook(OUTPUT, data_only=True)
ws_out = wb_out["Tổng hợp"]

employees_out = []
for r in range(2, ws_out.max_row + 1):
    code = ws_out.cell(row=r, column=2).value
    name = ws_out.cell(row=r, column=3).value
    if code and str(code).strip() and not str(code).strip().startswith("Mã"):
        employees_out.append((r, str(code).strip(), str(name).strip() if name else ""))

codes_ot = set(c[1] for c in employees_ot)
codes_out = set(c[2] for c in employees_out)

print(f"\n=== COMPARISON ===")
print(f"Overtime (2) sheet: {len(codes_ot)} employees")
print(f"Output LCNT: {len(codes_out)} employees")
print(f"\nIn both: {len(codes_ot & codes_out)}")
print(f"In OT only: {len(codes_ot - codes_out)}")
print(f"In output only: {len(codes_out - codes_ot)}")

# Show some samples to see if it's same employees but different format
print("\n=== Sample OT employees ===")
for c in sorted(codes_ot)[:20]:
    print(f"  {c}")
print("\n=== Sample output employees ===")
for c in sorted(codes_out)[:20]:
    print(f"  {c}")

# Check if maybe output uses 'OUT' sheet column 6+ which has sub-headers
print("\n=== Rerun: Check 'OUT' sheet structure deeper ===")
ws_out2 = wb["OUT"]
# Find header rows
for r in range(1, 12):
    print(f"R{r}: B='{ws_out2.cell(row=r, column=2).value}', C='{ws_out2.cell(row=r, column=3).value}', D='{ws_out2.cell(row=r, column=4).value}'")
