# -*- coding: utf-8 -*-
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, 'C:/CodeApp/HrP/appBCC')
os.chdir('C:/CodeApp/HrP/appBCC')

import openpyxl, openpyxl.utils as OU

wb_lcnt = openpyxl.load_workbook('C:/CodeApp/HrP/appBCC/docs/Actro/LCNT7.xlsx', data_only=True)
ws_lcnt = wb_lcnt.active
wb_bcc = openpyxl.load_workbook('C:/CodeApp/HrP/appBCC/docs/Actro/BCCActroT7.xlsx', data_only=True)
ws_bcc = wb_bcc.active

targets = {'A601010731': 15, 'A604011828': 44, 'A604011874': 45, 'A603011234': 78, 'A604011872': 80}

# Read LCNT7 KC:KK values for each employee
print('=== LCNT7 KC:KK (row from LCNT7 itself) ===')
for code, lcnt_row in sorted(targets.items(), key=lambda x: x[1]):
    kc = ws_lcnt.cell(row=lcnt_row, column=289).value
    kd = ws_lcnt.cell(row=lcnt_row, column=290).value
    ke = ws_lcnt.cell(row=lcnt_row, column=291).value
    kf = ws_lcnt.cell(row=lcnt_row, column=292).value
    kh = ws_lcnt.cell(row=lcnt_row, column=294).value
    ki = ws_lcnt.cell(row=lcnt_row, column=295).value
    kj = ws_lcnt.cell(row=lcnt_row, column=296).value
    kk = ws_lcnt.cell(row=lcnt_row, column=297).value
    kl = ws_lcnt.cell(row=lcnt_row, column=298).value  # CÔNG HC
    kt = ws_lcnt.cell(row=lcnt_row, column=306).value  # Chuyen can

    if kd and ke and ki:
        cong_hc = (kd + ke + ki) / 8
    else:
        cong_hc = None

    print(f'{code}:')
    print(f'  LCNT7: KC={kc} KD={kd} KE={ke} KF={kf} KH={kh} KI={ki} KJ={kj} KK={kk}')
    print(f'  LCNT7 KL (CONG HC) = {kl}')
    print(f'  Computed CONG HC = (KD+KE+KI)/8 = {cong_hc}')
    print(f'  LCNT7 ChuyenCan (KT) = {kt}')
    print()

# Also check BCC row for Vang Thin Thang to understand KD=104
print('=== Vang Thin Thang (A604011828) BCC raw data ===')
bcc_row = 42
print('BCC Row 42 all non-None cols 280-300:')
for c in range(280, 300):
    v = ws_bcc.cell(row=bcc_row, column=c).value
    if v is not None:
        print(f'  Col {OU.get_column_letter(c)}({c}): {v}')

print()
print('BCC Row 42 KC:KK formulas:')
wb_bcc_f = openpyxl.load_workbook('C:/CodeApp/HrP/appBCC/docs/Actro/BCCActroT7.xlsx', data_only=False)
ws_bcc_f = wb_bcc_f.active
for c in range(289, 298):
    v = ws_bcc_f.cell(row=bcc_row, column=c).value
    print(f'  Col {OU.get_column_letter(c)}({c}): {v}')
