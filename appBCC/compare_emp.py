# -*- coding: utf-8 -*-
"""Compare employees in input BCC vs output LCNT"""
import sys
sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)

import openpyxl

INPUT = r"C:\CodeApp\HrP\appBCC\docs\Actro\round2\BCCActroT7_OK.xlsx"
OUTPUT = r"C:\CodeApp\HrP\appBCC\docs\Actro\round2\LCNT_7_2026_BCC.xlsx"

print("="*100)
print("FILE INPUT: BCCActroT7_OK.xlsx")
print("="*100)

wb_in = openpyxl.load_workbook(INPUT, data_only=True)
print(f"Sheets: {wb_in.sheetnames}")
for sn in wb_in.sheetnames:
    ws = wb_in[sn]
    print(f"\n--- Sheet '{sn}' ---")
    print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")
    # First 3 rows
    for r in range(1, min(5, ws.max_row+1)):
        row_data = []
        for c in range(1, min(8, ws.max_column+1)):
            v = ws.cell(row=r, column=c).value
            row_data.append(str(v)[:20] if v is not None else "")
        print(f"  R{r}: {row_data}")

# Get employees list from input
print("\n" + "="*100)
print("EMPLOYEES IN INPUT BCC")
print("="*100)
ws_in = wb_in.active
employees_in = []
for r in range(3, ws_in.max_row + 1):
    code = ws_in.cell(row=r, column=2).value
    name = ws_in.cell(row=r, column=3).value
    if code and str(code).strip():
        employees_in.append((r, str(code).strip(), str(name).strip() if name else ""))
        
print(f"Total employees found: {len(employees_in)}")
for r, code, name in employees_in[:50]:
    print(f"  R{r}: {code} - {name}")

print("\n" + "="*100)
print("FILE OUTPUT: LCNT_7_2026_BCC.xlsx")
print("="*100)

wb_out = openpyxl.load_workbook(OUTPUT, data_only=True)
print(f"Sheets: {wb_out.sheetnames}")
for sn in wb_out.sheetnames:
    ws = wb_out[sn]
    print(f"\n--- Sheet '{sn}' ---")
    print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")
    for r in range(1, min(6, ws.max_row+1)):
        row_data = []
        for c in range(1, min(8, ws.max_column+1)):
            v = ws.cell(row=r, column=c).value
            row_data.append(str(v)[:20] if v is not None else "")
        print(f"  R{r}: {row_data}")

# Get employees from output sheet Tổng hợp
print("\n" + "="*100)
print("EMPLOYEES IN OUTPUT LCNT")
print("="*100)

employees_out = []
for sn in wb_out.sheetnames:
    ws = wb_out[sn]
    # Skip header rows
    for r in range(2, ws.max_row + 1):
        code = ws.cell(row=r, column=2).value
        name = ws.cell(row=r, column=3).value
        if code and str(code).strip() and not str(code).strip().startswith("Mã"):
            employees_out.append((sn, r, str(code).strip(), str(name).strip() if name else ""))
            
print(f"Total employees found in output: {len(employees_out)}")
for sn, r, code, name in employees_out[:50]:
    print(f"  [{sn}] R{r}: {code} - {name}")

# Compare codes
print("\n" + "="*100)
print("COMPARISON")
print("="*100)

codes_in = set(c[1] for c in employees_in)
codes_out = set(c[2] for c in employees_out)

print(f"\nIn input only: {len(codes_in - codes_out)}")
for c in sorted(codes_in - codes_out):
    print(f"  - {c}")

print(f"\nIn output only: {len(codes_out - codes_in)}")
for c in sorted(codes_out - codes_in):
    print(f"  - {c}")

print(f"\nIn both: {len(codes_in & codes_out)}")
