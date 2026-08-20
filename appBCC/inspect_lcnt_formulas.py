# -*- coding: utf-8 -*-
"""Re-check LCNT7 formulas and structure"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.utils import get_column_letter

# Look at row 11 (criteria) and row 13 (multipliers)
wb_f = openpyxl.load_workbook(r'C:\CodeApp\HrP\appBCC\docs\Actro\LCNT7.xlsx', data_only=False)
ws_f = wb_f.active

# Row 11 full
print("=== Row 11 (criteria - all cols) ===")
for col in range(280, 305):
    val = ws_f.cell(row=11, column=col).value
    if val is not None:
        print(f"  Col {col} ({get_column_letter(col)}): {repr(val)}")

# Row 13 full
print("\n=== Row 13 (multipliers - all cols) ===")
for col in range(280, 305):
    val = ws_f.cell(row=13, column=col).value
    if val is not None:
        print(f"  Col {col} ({get_column_letter(col)}): {val}")

# Look at formulas LCNT7 row 84 (A602010881)
print("\n=== Formulas at row 84 (A602010881) ===")
print("KD290 formula:")
print(f"  {ws_f.cell(row=84, column=290).value}")
print(f"KE291 formula:")
print(f"  {ws_f.cell(row=84, column=291).value}")
print(f"KI295 formula:")
print(f"  {ws_f.cell(row=84, column=295).value}")
print(f"KF292 formula:")
print(f"  {ws_f.cell(row=84, column=292).value}")

# Also row 15 (Ma Doan Chung) to confirm structure
print("\n=== Ma Doan Chung row 15 ===")
for col in range(290, 298):
    val = ws_f.cell(row=15, column=col).value
    print(f"  {get_column_letter(col)}{col}: {val}")

# Row 11 (criteria values that are used in SUMIFS)
# KD = SUMIFS(F:KC, $F$11:$KC$11, $JX$11)
# Let me check what is in JX11
print("\n=== Reference cells ===")
print(f"JX11: {ws_f.cell(row=11, column=286).value}")
print(f"JY11: {ws_f.cell(row=11, column=287).value}")
print(f"JZ11: {ws_f.cell(row=11, column=286).value}")  # JZ
print(f"JW11: {ws_f.cell(row=11, column=283).value}")  # JW = 1.5
print(f"KA11: {ws_f.cell(row=11, column=281).value}")  # KA = 200%
print(f"KB11: {ws_f.cell(row=11, column=282).value}")  # KB = 210%
print(f"HS11: {ws_f.cell(row=11, column=227).value}")
print(f"HT11: {ws_f.cell(row=11, column=229).value}")