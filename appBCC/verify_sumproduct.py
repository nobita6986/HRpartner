# -*- coding: utf-8 -*-
"""Verify HR sumproduct expected from daily breakdown"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.utils import get_column_letter

wb_bcc = openpyxl.load_workbook(r'C:\CodeApp\HrP\appBCC\docs\Actro\BCCActroT7.xlsx', data_only=True)
ws_bcc = wb_bcc.active

# Find day starts
day_cols = {}
for col in range(6, 290):
    val = ws_bcc.cell(row=8, column=col).value
    if val is not None and str(val).strip().isdigit():
        d = int(str(val).strip())
        if 1 <= d <= 31:
            day_cols[d] = col

# Sum each day based on HR SUMIFS criteria
# HR SUMIFS($F$11:$KC$11, criterion) where criterion values are in row 11
# Row 11 layout:
#   Col 'in' (F=6): 'In'
#   Col 'Out' (G=7): 'Out'
#   Col 'hours' (H=8): 'giờ'
#   Col JW (col 283): 1.5
#   Col JX: '8h~17h'
#   Col JY: '20h~22h'
#   Col JZ: 'Số giờ làm đêm'
#   Col KA: '200% ca dem/ 야간 잔업'
#   Col KB: '210% 야간 잔업'

# KD = sum of cells where row 11 = "8h~17h" (JX11 = '8h~17h')
# KE = sum of cells where row 11 = "20h~22h" (JY11 = '20h~22h')
# KF = sum of cells where row 11 = 1.5 (JW11 = 1.5)
# KH = sum of cells where row 11 = "200% ca dem" (KA11)
# KI = sum of cells where row 11 = "Số giờ đêm" (JZ11)
# KJ = sum of cells where row 11 = 2 (HS11 = 2)
# KK = sum of cells where row 11 = "Tổng" (HT11)

# KD = SUM cells in row 84 where ROW 11 COLUMN HEADER = "8h~17h"
# Wait, the formula is SUMIFS(F84:KC84, $F$11:$KC$11, $JX$11)
# This sums cells in row 84 where the corresponding cell in row 11 equals JX11 = "8h~17h"
# So it sums cells where the COLUMN is labeled "8h~17h" in row 11

# The cells in row 11 are the second-row sub-headers
# For day 26 (cols F-O): row 11 = [In, Out, giờ, 1.5, 8h~17h, 20h~22h, Số giờ đêm, 200% ca đêm, 210% CN, Tổng]
# For day 27: same pattern but starts at P
# For day 28 (CN): different: [In, Out, giờ, 2, 2.7, Tổng]
# For day 29: revert to standard 10-col pattern (F-O)

# So KD = sum all 8h~17h cells = h_8_17 from standard days
# KE = sum all 20h~22h cells
# KF = sum all 1.5 cells
# KH = sum all 200% ca đêm cells (night_200)
# KI = sum all Số giờ đêm cells (night_130)
# KJ = sum all 2 cells = from day 28 only (CN 200%)
# KK = sum all Tổng cells = O, Y, AE, AO, ...

# For A602010881 row 82:
row = 82

total_8h17 = 0  # KD
total_20h22 = 0  # KE
total_1_5 = 0  # KF
total_200 = 0  # KH (night_200)
total_night = 0  # KI (Số giờ đêm, 130%)
total_2 = 0  # KJ (cells labeled "2", from CN days)
total_tong = 0  # KK (Tổng)

for day_num in sorted(day_cols.keys()):
    start = day_cols[day_num]
    in_t = ws_bcc.cell(row=row, column=start).value
    
    # Read the row 11 sub-header for this day's cols
    sub_headers = []
    for offset in range(10):
        h = ws_bcc.cell(row=9, column=start+offset).value
        sub_headers.append(h)
    
    # And day-specific layout
    # Day 28: cols 26-31 (Z-AE) with subheaders: In, Out, giờ, 2, 2.7, Tổng
    # Standard: In, Out, giờ, 1.5, 8h~17h, 20h~22h, Số giờ đêm, 200% ca dem, 210% CN, Tổng
    
    for offset in range(10):
        if start + offset > 290:
            break
        sub = sub_headers[offset] if offset < len(sub_headers) else None
        val = ws_bcc.cell(row=row, column=start+offset).value
        
        if val is None or isinstance(val, str):
            continue
        
        # HR's SUMIFS criterion
        if sub == '8h~17h':
            total_8h17 += val
        elif sub == '20h~22h':
            total_20h22 += val
        elif sub == 1.5 or sub == '1.5':
            total_1_5 += val
        elif '200% ca dem' in str(sub):
            total_200 += val
        elif 'Số giờ' in str(sub) and 'đêm' in str(sub):
            total_night += val
        elif sub == 2 or str(sub) == '2':
            total_2 += val
        elif sub == 'Tổng':
            total_tong += val

print(f"=== For A602010881 (Ma A Sình) ===\n")
print(f"KD (8h~17h): {total_8h17}")
print(f"KE (20h~22h): {total_20h22}")
print(f"KF (1.5): {total_1_5}")
print(f"KH (200% đêm): {total_200}")
print(f"KI (Số giờ đêm): {total_night}")
print(f"KJ (2): {total_2}")
print(f"KK (Tổng): {total_tong}")

print(f"\n=== HR's BCC summary ===")
print(f"KD: {ws_bcc.cell(row=82, column=290).value}")
print(f"KE: {ws_bcc.cell(row=82, column=291).value}")
print(f"KF: {ws_bcc.cell(row=82, column=292).value}")
print(f"KH: {ws_bcc.cell(row=82, column=294).value}")
print(f"KI: {ws_bcc.cell(row=82, column=295).value}")
print(f"KJ: {ws_bcc.cell(row=82, column=296).value}")
print(f"KK: {ws_bcc.cell(row=82, column=297).value}")

print(f"\n=== SUMPRODUCT ===")
sumproduct = total_8h17 * 1 + total_20h22 * 1 + total_1_5 * 1.5 + total_200 * 2 + total_night * 1.3 + total_2 * 2 + total_tong * 2.7
print(f"Computed: {sumproduct}")
print(f"LCB = 6,000,000 / 26 / 8 = 28,846.15")
print(f"LƯƠNG = {28846.15 * sumproduct:,.0f}")