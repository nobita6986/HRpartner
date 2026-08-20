# -*- coding: utf-8 -*-
"""Check BCC vs LCNT7 for A602010881"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.utils import get_column_letter

wb_bcc = openpyxl.load_workbook(r'C:\CodeApp\HrP\appBCC\docs\Actro\BCCActroT7.xlsx', data_only=True)
ws_bcc = wb_bcc.active

wb_lcnt = openpyxl.load_workbook(r'C:\CodeApp\HrP\appBCC\docs\Actro\LCNT7.xlsx', data_only=True)
ws_lcnt = wb_lcnt.active

# Find A602010881 in both files
def find_employee(ws, code):
    for row in range(3, ws.max_row + 1):
        c = ws.cell(row=row, column=2).value
        if c and str(c).strip() == code:
            return row
    return None

code = 'A602010881'
bcc_row = find_employee(ws_bcc, code)
lcnt_row = find_employee(ws_lcnt, code)

print(f"BCC row: {bcc_row}, LCNT row: {lcnt_row}")

print(f"\n=== BCC SUMMARY (cols 289-297) ===")
for col in range(289, 298):
    val_bcc = ws_bcc.cell(row=bcc_row, column=col).value
    print(f"  {get_column_letter(col)} (col {col}): {val_bcc}")

print(f"\n=== LCNT7 SUMMARY (cols 289-297) ===")
for col in range(289, 298):
    val_lcnt = ws_lcnt.cell(row=lcnt_row, column=col).value
    print(f"  {get_column_letter(col)} (col {col}): {val_lcnt}")

# Now compare side by side
print(f"\n=== COMPARISON (BCC vs LCNT7) ===")
keys = {
    289: 'KC',
    290: 'KD (Số giờ ban ngày)',
    291: 'KE',
    292: 'KF (Shift day/주간)',
    294: 'KH (Night day/야간)',
    295: 'KI',
    296: 'KJ (Sunday)',
    297: 'KK',
}
for col, key in keys.items():
    val_bcc = ws_bcc.cell(row=bcc_row, column=col).value
    val_lcnt = ws_lcnt.cell(row=lcnt_row, column=col).value
    match = "✓" if val_bcc == val_lcnt else "✗ DIFF"
    print(f"  {key} ({get_column_letter(col)}{col}): BCC={val_bcc}, LCNT={val_lcnt} {match}")

# Check what BCC daily shows
print(f"\n=== BCC daily breakdown (col indices) ===")
# Day 26 starts at col 6 (F). Layout: in, out, giờ, 1.5, 8h17, 20h22, đêm, 200% đêm, 210% CN
# So day 26 col offset: in=5, out=6, hours=7, 1.5x=8, 8h17=9, 20h22=10, đêm=11, 200đêm=12, 210CN=13 (pandas 0-based)
print(f"  Day 26 starts at col F (index 5)")
print(f"  F: In, G: Out, H: hours, I: 1.5x OT, J: 8h~17h, K: 20h~22h, L: Night, M: 200% đêm, N: 210% CN")

# Sum up daily OT day
print(f"\n=== BCC Daily OT day (col I = 1.5x) ===")
total_ot_day = 0
for day_idx in range(26, 32):  # days 26-31
    day_num = day_idx
    col_offset = 5 + (day_idx - 26) * 9  # day 26 starts at col 5
    if day_idx > 25:
        col_offset = 5 + (day_idx - 26) * 9
    # This is too complex, let me just look at day 26 directly
print("Looking at col offsets for day 26:")