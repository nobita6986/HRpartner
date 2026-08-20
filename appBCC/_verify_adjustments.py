import sys

sys.path.insert(0, r"C:\CodeApp\HrP\appBCC")

import openpyxl
from pathlib import Path

from adjustments_template import build_adjustment_template, DEFAULT_ADJUSTMENT_PAIRS
from adjustments import parse_adjustment_file, apply_adjustments_to_payroll

sample_path = Path(r"C:\CodeApp\HrP\appBCC\_adjustment_sample.xlsx")
build_adjustment_template(sample_path)

workbook = openpyxl.load_workbook(sample_path)
sheet = workbook["Danh sách khoản"]
sheet.cell(row=2, column=1, value="A601010731")
sheet.cell(row=2, column=2, value="Ma Doãn Chung")
sheet.cell(row=2, column=3, value=200000)
sheet.cell(row=2, column=4, value=50000)
sheet.cell(row=2, column=5, value=150000)
sheet.cell(row=2, column=7, value=100000)
sheet.cell(row=3, column=1, value="A600000000")
sheet.cell(row=3, column=2, value="Test NV")
sheet.cell(row=3, column=3, value=300000)
sheet.cell(row=3, column=4, value=0)
workbook.save(sample_path)
workbook.close()

data = parse_adjustment_file(sample_path)
print("PARSED:", data)
assert "A601010731" in data
assert len(data["A601010731"]) == 4
assert data["A601010731"][0].kind == "PLUS"
assert data["A601010731"][0].name == "Thưởng chuyên cần"
assert data["A601010731"][1].kind == "MINUS"
assert data["A601010731"][3].name == "Phụ cấp đặc biệt"

workbook = openpyxl.load_workbook(sample_path)
sheet = workbook["Danh sách khoản"]
sheet.cell(row=1, column=3, value="Thưởng chuyên cần Actro")
sheet.cell(row=1, column=4, value="Trừ đi muộn Actro")
workbook.save(sample_path)
workbook.close()

data = parse_adjustment_file(sample_path)
print("RENAMED:", data)
assert data["A601010731"][0].name == "Thưởng chuyên cần Actro"
assert data["A601010731"][1].name == "Trừ đi muộn Actro"

sample_payroll = apply_adjustments_to_payroll(
    {
        "salaryItems": [],
        "allowances": [],
        "deductions": [],
        "summary": {"salaryNormal": 0, "thanhToan": 0, "totalDeduction": 0, "netIncome": 0},
    },
    data["A601010731"],
)
print("NET:", sample_payroll["summary"]["netIncome"])
assert any(item["name"] == "Thưởng chuyên cần Actro" for item in sample_payroll["salaryItems"])
assert any(item["name"] == "Trừ đi muộn Actro" for item in sample_payroll["deductions"])
assert sample_payroll["adjustments"][0]["kind"] == "PLUS"

print("ALL_OK")
