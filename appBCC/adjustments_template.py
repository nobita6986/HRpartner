"""Tạo file mẫu cho phép người dùng nhập các khoản cộng/trừ linh hoạt theo nhân viên.

Cấu trúc file:
- Sheet ``Danh sách khoản``: tiêu đề từng khoản, dạng cặp cộng/trừ theo vị trí cột.
  Mỗi cặp gồm 2 cột: cột lẻ là CỘNG (tiền thưởng / phụ cấp đặc biệt), cột chẵn là TRỪ
  (phạt / cấn trừ). Header được tô màu xanh cho cộng, màu đỏ cho trừ.
- Sheet ``Hướng dẫn``: chỉ dẫn chi tiết để người dùng tự đổi tên cột tuỳ ý.

Khi đọc lại, app lấy header của từng cột (kể cả khi người dùng đổi tên) và gán
theo vị trí: cột lẻ -> cộng, cột chẵn -> trừ. Tên cột mới sẽ được dùng làm tên
khoản trong file Excel chuẩn hoá.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ADJUSTMENT_TEMPLATE_FILENAME = "MAU_CAP_NHAP_KHOAN_THUONG_PHAT.xlsx"

# Cặp khoản mặc định: (tên cộng, tên trừ). Có thể mở rộng bằng cách thêm cặp.
DEFAULT_ADJUSTMENT_PAIRS = (
    ("Thưởng chuyên cần", "Phạt chuyên cần"),
    ("Thưởng năng suất", "Phạt chậm tiến độ"),
    ("Phụ cấp đặc biệt", "Cấn trừ khác"),
    ("Thưởng khác", "Phạt khác"),
)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, color="FFFFFF", bold=True)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F4E78")
NOTE_FONT = Font(name="Arial", size=10, italic=True, color="595959")
SAMPLE_FONT = Font(name="Arial", size=10)
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def build_adjustment_template(output_path: str | Path) -> Path:
    """Tạo file mẫu cộng/trừ và trả về đường dẫn tuyệt đối."""
    path = Path(output_path).resolve()
    workbook = openpyxl.Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Danh sách khoản"

    headers = ["Mã NV", "Họ tên"]
    for plus_name, minus_name in DEFAULT_ADJUSTMENT_PAIRS:
        headers.append(plus_name)
        headers.append(minus_name)
    data_sheet.append(headers)

    data_sheet.row_dimensions[1].height = 32
    for col_idx, header in enumerate(headers, start=1):
        cell = data_sheet.cell(row=1, column=col_idx, value=header)
        cell.alignment = CENTER
        cell.font = HEADER_FONT
        cell.border = BORDER
        if header == "Mã NV":
            cell.fill = HEADER_FILL
        elif header == "Họ tên":
            cell.fill = HEADER_FILL
        else:
            pair_index = (col_idx - 3) // 2
            if 0 <= pair_index < len(DEFAULT_ADJUSTMENT_PAIRS):
                cell.fill = GREEN_FILL if (col_idx - 3) % 2 == 0 else RED_FILL
            cell.font = Font(name="Arial", size=11, bold=True)

    sample_rows = [
        ("A601010731", "Ma Doãn Chung"),
        ("A603011167", "Vừ A Cù"),
    ]
    for row in sample_rows:
        data_sheet.append(row)
    for row in data_sheet.iter_rows(min_row=2, max_row=1 + len(sample_rows)):
        for cell in row:
            cell.font = SAMPLE_FONT
            cell.border = BORDER
            cell.alignment = LEFT

    widths = {"A": 14, "B": 22}
    for col_idx in range(3, len(headers) + 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        widths[column_letter] = 18
    for column_letter, width in widths.items():
        data_sheet.column_dimensions[column_letter].width = width

    data_sheet.freeze_panes = "C2"

    guide_sheet = workbook.create_sheet("Hướng dẫn")
    instructions = [
        ("Cách dùng file mẫu", "bold"),
        ("1. Copy file mẫu này, đổi tên theo kỳ lương (ví dụ: BONUS_PENALTY_T6_2026.xlsx).", "normal"),
        ("2. Sheet 'Danh sách khoản' chứa các cặp cộng/trừ. Mỗi cặp gồm 2 cột liên tiếp: cột lẻ (xanh) là CỘNG tiền, cột chẵn (đỏ) là TRỪ tiền.", "normal"),
        ("3. Bạn có thể đổi tên cột tuỳ ý — tên mới sẽ xuất hiện đúng nội dung trong file Excel chuẩn hoá lương.", "normal"),
        ("4. Nhập mã nhân viên đúng theo BCC (cột A). Để trống họ tên, app sẽ tự tra theo mã.", "normal"),
        ("5. Chỉ cần điền vào các khoản thực sự phát sinh. Khoản nào để trống sẽ bị bỏ qua, không ảnh hưởng lương.", "normal"),
        ("6. Upload file đã điền từ nút 'Tải danh sách cộng/trừ' trong app sau khi đã tính lương.", "normal"),
        ("", "normal"),
        ("Quy tắc phân biệt cộng/trừ", "bold"),
        ("- Cột lẻ (kể từ cột C — cột đầu tiên sau Mã NV/Họ tên): app hiểu là khoản CỘNG (tiền thưởng, phụ cấp đặc biệt, ...).", "normal"),
        ("- Cột chẵn: app hiểu là khoản TRỪ (phạt, cấn trừ, ...).", "normal"),
        ("- Có thể thêm/bớt cặp cột tuỳ ý, miễn đúng thứ tự CỘNG -> TRỪ -> CỘNG -> TRỪ ...", "normal"),
        ("", "normal"),
        ("Sau khi áp dụng", "bold"),
        ("- Các khoản cộng sẽ xuất hiện trong mảng 'salaryItems' và tăng tổng lương.", "normal"),
        ("- Các khoản trừ sẽ xuất hiện trong mảng 'deductions' và giảm tổng lương.", "normal"),
        ("- Tổng thu nhập = Lương cơ bản + Phụ cấp + Cộng − Trừ.", "normal"),
    ]
    guide_sheet.column_dimensions["A"].width = 110
    for row_idx, (text, style) in enumerate(instructions, start=1):
        cell = guide_sheet.cell(row=row_idx, column=1, value=text)
        if style == "bold":
            cell.font = TITLE_FONT
        elif text.startswith("-"):
            cell.font = NOTE_FONT
        else:
            cell.font = SAMPLE_FONT
        cell.alignment = LEFT
        guide_sheet.row_dimensions[row_idx].height = 32 if style == "bold" else 20

    workbook.save(path)
    return path


if __name__ == "__main__":
    output = build_adjustment_template(ADJUSTMENT_TEMPLATE_FILENAME)
    print(f"Đã tạo file mẫu: {output}")
