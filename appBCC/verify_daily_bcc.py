# -*- coding: utf-8 -*-
"""So sánh BCC daily times với HR cho Ma A Sình"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.utils import get_column_letter

wb_bcc = openpyxl.load_workbook(r'C:\CodeApp\HrP\appBCC\docs\Actro\BCCActroT7.xlsx', data_only=True)
ws_bcc = wb_bcc.active

# Find Ma A Sình in BCC row 82
code = 'A602010881'
print(f"=== BCC Daily for {code} ===")
print(f"{'Day':<5} | {'In':<7} | {'Out':<7} | {'Hours':<6} | {'OT1.5':<6} | {'8h17':<6} | {'20h22':<6} | {'Đêm130':<7} | {'200%đêm':<8} | {'210CN':<6}")
print("-" * 110)

# Day 26 starts at col F=6 (1-based). 
# But the structure changed at day 28 (cols 26-31).
# After day 28, layout reverts to standard
# Let me look at row 8 headers to find day starts

day_cols = {}
for col in range(6, 290):
    val = ws_bcc.cell(row=8, column=col).value
    if val is not None and str(val).strip().isdigit():
        d = int(str(val).strip())
        if 1 <= d <= 31:
            day_cols[d] = col

print(f"Day columns found: {sorted(day_cols.keys())[:10]}...{sorted(day_cols.keys())[-5:]}")
print(f"Day 26 starts at col: {day_cols.get(26)}")
print(f"Day 28 starts at col: {day_cols.get(28)}")
print(f"Day 29 starts at col: {day_cols.get(29)}")

# Show all days for this employee
for day_num in sorted(day_cols.keys()):
    start = day_cols[day_num]
    in_t = ws_bcc.cell(row=82, column=start).value
    out_t = ws_bcc.cell(row=82, column=start+1).value
    hours = ws_bcc.cell(row=82, column=start+2).value
    
    # Check layout based on start col
    # Day 26 (start 6): in=F, out=G, hours=H, ot1.5=I, 8h17=J, 20h22=K, đêm=L, 200%đêm=M, 210CN=N, tổng=O
    # Day 28 (start 26): different layout
    if start == 26:
        # Day 28 layout: AC=2 (200% CN), AD=2.7 (270% CN), no 1.5, no đêm
        ot_1_5 = '-'
        h_8_17 = '-'
        h_20_22 = '-'
        night_130 = '-'
        night_200 = ws_bcc.cell(row=82, column=start+3).value  # AC
        cn_210 = ws_bcc.cell(row=82, column=start+4).value  # AD
    else:
        ot_1_5 = ws_bcc.cell(row=82, column=start+3).value
        h_8_17 = ws_bcc.cell(row=82, column=start+4).value
        h_20_22 = ws_bcc.cell(row=82, column=start+5).value
        night_130 = ws_bcc.cell(row=82, column=start+6).value
        night_200 = ws_bcc.cell(row=82, column=start+7).value
        cn_210 = ws_bcc.cell(row=82, column=start+8).value
    
    in_str = str(in_t).split(' ')[-1] if in_t else ''
    out_str = str(out_t).split(' ')[-1] if out_t else ''
    
    if any([in_t, out_t, hours, ot_1_5, night_130, night_200, cn_210]):
        print(f"{day_num:<5} | {in_str:<7} | {out_str:<7} | {str(hours or 0):<6} | {str(ot_1_5 or 0):<6} | {str(h_8_17 or 0):<6} | {str(h_20_22 or 0):<6} | {str(night_130 or 0):<7} | {str(night_200 or 0):<8} | {str(cn_210 or 0):<6}")

# Now manually compute HR values
print(f"\n=== Manual SUMPRODUCT (matching HR) ===")
total_8_17 = 0  # KD
total_20_22 = 0  # KE
total_ot_1_5 = 0  # KF
total_out = 0  # KH (Night day)
total_night_130 = 0  # KI
total_200_cn_or_2 = 0  # KJ
total_tong = 0  # KK

for day_num in sorted(day_cols.keys()):
    start = day_cols[day_num]
    in_t = ws_bcc.cell(row=82, column=start).value
    out_t = ws_bcc.cell(row=82, column=start+1).value
    hours = ws_bcc.cell(row=82, column=start+2).value
    
    if start == 26:  # Day 28 special layout
        # No 1.5, 8h17, 20h22, đêm, 200%đêm in this day type
        # AC=200% CN, AD=270% CN
        night_200 = ws_bcc.cell(row=82, column=start+3).value or 0
        cn_210 = ws_bcc.cell(row=82, column=start+4).value or 0
        # In sum:
        # JX (8h17) = ? - depends on what HR's SUMIFS returns when column doesn't exist
        # For day 28, JX returns nothing, KE also nothing
        # KJ (HS11=2) - sums cells where =2 - matches AC col
        # KK (HT11=Tổng) - matches the O column (= AE for day 28)
        tong = ws_bcc.cell(row=82, column=start+5).value or 0  # AE
        if in_t:  # Day has work
            total_200_cn_or_2 += cn_210  # KJ sums CN=210% hours
            total_tong += tong  # KK sums totals
    else:
        ot_1_5 = ws_bcc.cell(row=82, column=start+3).value or 0
        h_8_17 = ws_bcc.cell(row=82, column=start+4).value or 0
        h_20_22 = ws_bcc.cell(row=82, column=start+5).value or 0
        night_130 = ws_bcc.cell(row=82, column=start+6).value or 0
        night_200 = ws_bcc.cell(row=82, column=start+7).value or 0
        cn_210 = ws_bcc.cell(row=82, column=start+8).value or 0
        tong = ws_bcc.cell(row=82, column=start+9).value or 0  # O/AE/O...
        
        if in_t:
            total_8_17 += h_8_17
            total_20_22 += h_20_22
            total_ot_1_5 += ot_1_5
            total_out += out_t if str(out_t).strip() else 0  # KH = sum of "Out" cells = ?
            # Wait, KH = SUMIFS where criteria = KA11 = "Out". So it sums all cells where row 11 = "Out"
            # "Out" is column G for normal day (col G), Q, AA, etc.
            # Wait but Out is a TIME value, it can't be summed numerically
            # Let me re-check...

print(f"  Total 8h17 (KD expected): {total_8_17}")
print(f"  Total 20h22 (KE expected): {total_20_22}")
print(f"  Total OT 1.5 (KF): {total_ot_1_5}")
print(f"  Total Night 130 (KI expected): {total_night_130}")

print(f"\n  HR values:")
print(f"  KD={ws_bcc.cell(row=82, column=290).value}")
print(f"  KE={ws_bcc.cell(row=82, column=291).value}")
print(f"  KF={ws_bcc.cell(row=82, column=292).value}")
print(f"  KH={ws_bcc.cell(row=82, column=294).value}")
print(f"  KI={ws_bcc.cell(row=82, column=295).value}")