# -*- coding: utf-8 -*-
"""
Export Manager - Xuất file Excel chuẩn hoá giống LCNT7 + Push DB

Cấu trúc payrollData:
- salaryItems: Lương cơ bản + OT breakdown
- allowances: Phụ cấp (chuyên cần, đời sống, thâm niên, nhà ở, soi kính)
- deductions: Khấu trừ (trừ ứng)
- summary: Tổng hợp (salaryNormal, thanhToan, totalDeduction, netIncome)

Actro column mapping (KS:LA):
  KS  : Lương giờ
  KT  : Chuyên cần HR
  KU  : Phụ cấp soi kính
  KV  : Phụ cấp đời sống
  KW  : Phụ cấp nhà ở (INPUT)
  KX  : Thâm niên
  KY  : Thanh toán = SUM(KT:KX)
  KZ  : Trừ ứng (INPUT)
  LA  : Thực nhận = ROUNDDOWN(KY-KZ, -3)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import calendar


class ExportManager:
    """Quản lý export dữ liệu tính lương"""

    def __init__(self, parsed_data: list, project_name: str, period_month: int, period_year: int):
        self.data = parsed_data
        self.project_name = project_name
        self.period_month = period_month
        self.period_year = period_year

        # Style definitions
        self.header_font = Font(name='Arial', size=10, bold=True)
        self.header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        self.header_fill_orange = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
        self.header_fill_ot = PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")
        self.data_font = Font(name='Arial', size=9)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')

    def export_to_excel(self, output_path: str) -> dict:
        """Xuất file Excel đầy đủ dữ liệu để đối chiếu"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self._create_summary_sheet(wb)
        self._create_payroll_detail_sheet(wb)
        self._create_daily_sheet(wb)
        self._create_ot_detail_sheet(wb)

        db_payload = self._prepare_db_payload()
        wb.save(output_path)

        return {
            "success": True,
            "sheets_created": ["Tổng hợp", "Chi tiết Lương", "Công hàng ngày", "Chi tiết OT"],
            "db_payload": db_payload,
            "total_employees": len(self.data)
        }

    def _get_emp_info(self, emp: dict) -> dict:
        """Extract employee info và payroll data"""
        payroll = emp.get('payrollData') or emp.get('payroll') or {}

        # Summary
        summary = payroll.get('summary', {}) if isinstance(payroll, dict) else {}

        # Extract từ salaryItems
        salary_items = payroll.get('salaryItems', []) if isinstance(payroll, dict) else []
        allowances = payroll.get('allowances', []) if isinstance(payroll, dict) else []
        deductions = payroll.get('deductions', []) if isinstance(payroll, dict) else []

        # OT breakdown
        ot_day = 0.0
        ot_night = 0.0
        ot_sunday = 0.0
        ot_holiday = 0.0

        for item in salary_items:
            name = item.get('name', '')
            qty = item.get('qty', 0) or 0
            if 'OT ngày' in name:
                ot_day = qty
            elif 'OT đêm' in name:
                ot_night = qty
            elif 'OT chủ nhật' in name:
                ot_sunday = qty
            elif 'OT ngày lễ' in name:
                ot_holiday = qty

        # Allowances breakdown — Actro only: KT:KX
        chuyen_can = 0.0
        soi_kinh = 0.0
        doi_song = 0.0
        nha_o = 0.0
        tham_nien = 0.0
        other_allowances = 0.0
        total_allowance = 0.0

        for item in allowances:
            name = item.get('name', '')
            total = item.get('total', 0) or 0
            total_allowance += total
            normalized_name = name.casefold()
            if 'chuyên cần' in normalized_name:
                chuyen_can = total
            elif 'soi kính' in normalized_name:
                soi_kinh = total
            elif 'đời sống' in normalized_name:
                doi_song = total
            elif 'nhà ở' in normalized_name:
                nha_o = total
            elif 'thâm niên' in normalized_name:
                tham_nien = total
            else:
                other_allowances += total

        # Deductions
        tru_ung = 0.0
        other_deductions = 0.0

        for item in deductions:
            name = item.get('name', '')
            total = item.get('total', 0) or 0
            if 'trừ ứng' in name.casefold():
                tru_ung = total
            else:
                other_deductions += total

        return {
            'employee_code': emp.get('employeeCode', ''),
            'full_name': emp.get('fullName', ''),
            'join_date': emp.get('joinDate', ''),
            'daily_data': emp.get('dailyData', []),
            'raw': emp,

            # Tổng công
            'total_work_days': emp.get('totalWorkDays', 0),
            'ot_hours': emp.get('otHours', 0),
            'absent_days': emp.get('absentDays', 0),

            # OT breakdown
            'ot_day': ot_day,
            'ot_night': ot_night,
            'ot_sunday': ot_sunday,
            'ot_holiday': ot_holiday,

            # Lương từ summary
            'salary_normal': summary.get('salaryNormal', 0),
            'salary_ot': summary.get('salaryOT', 0),
            'total_salary': summary.get('totalSalary', 0),

            # Actro allowances
            'chuyen_can': chuyen_can,
            'soi_kinh': soi_kinh,
            'doi_song': doi_song,
            'nha_o': nha_o,
            'tham_nien': tham_nien,
            'other_allowances': other_allowances,
            'total_allowance': total_allowance,

            # KY = SUM(KS:KX) = Lương giờ + Tổng phụ cấp
            'thanh_toan': summary.get('salaryNormal', 0) + total_allowance,

            # LA = ROUNDDOWN(KY-KZ, -3)
            'net_income': summary.get('netIncome', 0),
            'rounding_delta': summary.get('roundingDelta', 0),

            # Deductions
            'tru_ung': tru_ung,
            'other_deductions': other_deductions,
            'total_deduction': summary.get('totalDeduction', 0),
        }

    def _create_summary_sheet(self, wb):
        """Sheet Tổng hợp - giống LCNT7"""
        ws = wb.create_sheet("Tổng hợp", 0)

        # Header
        ws.cell(row=1, column=1, value="BẢNG CHẤM CÔNG NHÂN LỰC THỜI VỤ")
        ws.cell(row=1, column=1).font = Font(name='Arial', size=14, bold=True)
        ws.merge_cells('A1:R1')

        ws.cell(row=2, column=1, value=f"CÔNG TY {self.project_name}")
        ws.cell(row=2, column=1).font = Font(name='Arial', size=11, bold=True)
        ws.merge_cells('A2:R2')

        ws.cell(row=3, column=1, value=f"Tháng {self.period_month}/{self.period_year}")
        ws.cell(row=3, column=1).font = Font(name='Arial', size=10, italic=True)
        ws.merge_cells('A3:R3')

        # Headers row 5 — Actro KS:LA mapping
        headers = [
            ("STT", 5),
            ("Mã thẻ", 15),
            ("Họ và tên", 25),
            ("Ngày vào", 12),
            ("Nhà máy", 15),
            ("Số ngày công", 12),
            ("Số giờ OT", 10),
            ("Lương giờ (KS)", 15),
            ("Chuyên cần HR (KT)", 16),
            ("Phụ cấp soi kính (KU)", 18),
            ("Phụ cấp đời sống (KV)", 18),
            ("Phụ cấp nhà ở (KW)", 16),
            ("Thâm niên (KX)", 14),
            ("Phụ cấp khác", 14),
            ("Tổng phụ cấp", 15),
            ("Thanh toán (KY)", 15),
            ("Trừ ứng (KZ)", 12),
            ("Thực nhận (LA)", 15),
        ]

        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # Data rows
        for idx, emp in enumerate(self.data):
            row = 6 + idx
            info = self._get_emp_info(emp)

            ws.cell(row=row, column=1, value=idx + 1)
            ws.cell(row=row, column=2, value=info['employee_code'])
            ws.cell(row=row, column=3, value=info['full_name'])
            ws.cell(row=row, column=4, value=info['join_date'])
            ws.cell(row=row, column=5, value=self.project_name)
            ws.cell(row=row, column=6, value=info['total_work_days'])
            ws.cell(row=row, column=7, value=info['ot_hours'])
            ws.cell(row=row, column=8, value=info['salary_normal'])
            ws.cell(row=row, column=9, value=info['chuyen_can'])
            ws.cell(row=row, column=10, value=info['soi_kinh'])
            ws.cell(row=row, column=11, value=info['doi_song'])
            ws.cell(row=row, column=12, value=info['nha_o'])
            ws.cell(row=row, column=13, value=info['tham_nien'])
            ws.cell(row=row, column=14, value=info['other_allowances'])
            ws.cell(row=row, column=15, value=info['total_allowance'])
            ws.cell(row=row, column=16, value=info['thanh_toan'])
            ws.cell(row=row, column=17, value=info['tru_ung'])
            ws.cell(row=row, column=18, value=info['net_income'])

            # Format
            for col in range(1, 19):
                cell = ws.cell(row=row, column=col)
                cell.font = self.data_font
                cell.border = self.border
                if col >= 6:
                    cell.number_format = '#,##0'

    def _create_payroll_detail_sheet(self, wb):
        """Sheet Chi tiết Lương"""
        ws = wb.create_sheet("Chi tiết Lương")

        headers = [
            ("STT", 5),
            ("Mã NV", 12),
            ("Họ tên", 25),
            ("Ngày vào", 12),
            ("Lương giờ (KS)", 15),
            ("OT ngày (h)", 10),
            ("OT đêm (h)", 10),
            ("OT CN/Lễ (h)", 12),
            ("Tiền OT", 15),
            ("Chuyên cần HR (KT)", 16),
            ("Phụ cấp soi kính (KU)", 18),
            ("Phụ cấp đời sống (KV)", 18),
            ("Phụ cấp nhà ở (KW)", 16),
            ("Thâm niên (KX)", 14),
            ("Phụ cấp khác", 14),
            ("Tổng phụ cấp", 15),
            ("Thanh toán (KY)", 15),
            ("Trừ ứng (KZ)", 12),
            ("Khấu trừ khác", 14),
            ("Thực nhận (LA)", 15),
        ]

        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        for idx, emp in enumerate(self.data):
            row = 2 + idx
            info = self._get_emp_info(emp)

            ws.cell(row=row, column=1, value=idx + 1)
            ws.cell(row=row, column=2, value=info['employee_code'])
            ws.cell(row=row, column=3, value=info['full_name'])
            ws.cell(row=row, column=4, value=info['join_date'])
            ws.cell(row=row, column=5, value=info['salary_normal'])
            ws.cell(row=row, column=6, value=info['ot_day'])
            ws.cell(row=row, column=7, value=info['ot_night'])
            ws.cell(row=row, column=8, value=info['ot_sunday'])
            ws.cell(row=row, column=9, value=info['salary_ot'])
            ws.cell(row=row, column=10, value=info['chuyen_can'])
            ws.cell(row=row, column=11, value=info['soi_kinh'])
            ws.cell(row=row, column=12, value=info['doi_song'])
            ws.cell(row=row, column=13, value=info['nha_o'])
            ws.cell(row=row, column=14, value=info['tham_nien'])
            ws.cell(row=row, column=15, value=info['other_allowances'])
            ws.cell(row=row, column=16, value=info['total_allowance'])
            ws.cell(row=row, column=17, value=info['thanh_toan'])
            ws.cell(row=row, column=18, value=info['tru_ung'])
            ws.cell(row=row, column=19, value=info['other_deductions'])
            ws.cell(row=row, column=20, value=info['net_income'])

            # Format
            for col in range(1, 21):
                cell = ws.cell(row=row, column=col)
                cell.font = self.data_font
                cell.border = self.border
                if col >= 5:
                    cell.number_format = '#,##0'

    def _create_daily_sheet(self, wb):
        """Sheet Công hàng ngày"""
        ws = wb.create_sheet("Công hàng ngày")

        days_in_month = calendar.monthrange(self.period_year, self.period_month)[1]
        weekday_names = ["T2", "T3", "T4", "T5", "T6", "T7", "CN"]
        import datetime

        headers = [("STT", 5), ("Mã NV", 12), ("Họ tên", 25)]
        for day in range(1, days_in_month + 1):
            wd = weekday_names[datetime.date(self.period_year, self.period_month, day).weekday()]
            headers.append((f"{day}/{wd}", 8))
        headers.append(("Tổng", 10))

        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        for idx, emp in enumerate(self.data):
            row = 2 + idx
            info = self._get_emp_info(emp)
            daily_data = info['daily_data']

            ws.cell(row=row, column=1, value=idx + 1)
            ws.cell(row=row, column=2, value=info['employee_code'])
            ws.cell(row=row, column=3, value=info['full_name'])

            total_hours = 0
            col = 4
            for day in range(1, days_in_month + 1):
                date_suffix = f"-{day:02d}"
                day_info = next((d for d in daily_data if d.get("date", "").endswith(date_suffix)), {})

                hours = day_info.get('regularHours', 0) or day_info.get('hours', 0) or 0
                ot = day_info.get('ot', 0) or 0
                day_total = hours + ot

                ws.cell(row=row, column=col, value=day_total)
                total_hours += day_total
                col += 1

            ws.cell(row=row, column=col, value=total_hours)
            ws.cell(row=row, column=col).font = Font(name='Arial', size=9, bold=True)

            # Format
            for c in range(1, col + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = self.data_font
                cell.border = self.border
                if c >= 4:
                    cell.number_format = '0.0'

    def _create_ot_detail_sheet(self, wb):
        """Sheet Chi tiết OT"""
        ws = wb.create_sheet("Chi tiết OT")

        headers = [
            ("STT", 5),
            ("Mã NV", 12),
            ("Họ tên", 25),
            ("Tổng OT (h)", 10),
            ("OT ngày 150% (h)", 14),
            ("OT đêm 200% (h)", 14),
            ("OT CN 200% (h)", 14),
            ("OT lễ 300% (h)", 14),
            ("Tiền OT ngày", 15),
            ("Tiền OT đêm", 15),
            ("Tiền OT CN", 15),
            ("Tiền OT lễ", 15),
            ("Tổng tiền OT", 15),
        ]

        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill_ot
            cell.alignment = self.center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        for idx, emp in enumerate(self.data):
            row = 2 + idx
            info = self._get_emp_info(emp)

            # Calculate OT pay breakdown
            salary_items = emp.get('payrollData', {}).get('salaryItems', []) if emp.get('payrollData') else []
            ot_day_pay = 0.0
            ot_night_pay = 0.0
            ot_sunday_pay = 0.0
            ot_holiday_pay = 0.0

            for item in salary_items:
                name = item.get('name', '')
                total = item.get('total', 0) or 0
                if 'OT ngày' in name:
                    ot_day_pay = total
                elif 'OT đêm' in name:
                    ot_night_pay = total
                elif 'OT chủ nhật' in name:
                    ot_sunday_pay = total
                elif 'OT ngày lễ' in name:
                    ot_holiday_pay = total

            ws.cell(row=row, column=1, value=idx + 1)
            ws.cell(row=row, column=2, value=info['employee_code'])
            ws.cell(row=row, column=3, value=info['full_name'])
            ws.cell(row=row, column=4, value=info['ot_day'] + info['ot_night'] + info['ot_sunday'] + info['ot_holiday'])
            ws.cell(row=row, column=5, value=info['ot_day'])
            ws.cell(row=row, column=6, value=info['ot_night'])
            ws.cell(row=row, column=7, value=info['ot_sunday'])
            ws.cell(row=row, column=8, value=info['ot_holiday'])
            ws.cell(row=row, column=9, value=ot_day_pay)
            ws.cell(row=row, column=10, value=ot_night_pay)
            ws.cell(row=row, column=11, value=ot_sunday_pay)
            ws.cell(row=row, column=12, value=ot_holiday_pay)
            ws.cell(row=row, column=13, value=info['salary_ot'])

            # Format
            for col in range(1, 14):
                cell = ws.cell(row=row, column=col)
                cell.font = self.data_font
                cell.border = self.border
                if col >= 4:
                    cell.number_format = '#,##0'

    def _prepare_db_payload(self) -> dict:
        """Chuẩn bị payload để push lên DB"""
        timesheets = []
        timesheet_lines = []

        for emp in self.data:
            info = self._get_emp_info(emp)
            daily_data = info['daily_data']
            payroll = emp.get('payrollData') or {}

            # Timesheet record
            timesheet = {
                "employeeCode": info['employee_code'],
                "fullName": info['full_name'],
                "project": self.project_name,
                "periodMonth": self.period_month,
                "periodYear": self.period_year,
                "totalWorkDays": float(info['total_work_days']),
                "otHours": float(info['ot_hours']),
                "absentDays": float(info['absent_days']),
                "dailyData": daily_data,
                "payrollData": payroll,
                "totalIncome": float(info['net_income']),
            }
            timesheets.append(timesheet)

            # Daily lines
            for day_info in daily_data:
                date_str = day_info.get("date", "")
                if not date_str:
                    continue

                parts = date_str.split("-")
                if len(parts) >= 3:
                    day = int(parts[2])
                else:
                    continue

                work_date = f"{self.period_year}-{self.period_month:02d}-{day:02d}"

                line = {
                    "employeeCode": info['employee_code'],
                    "project": self.project_name,
                    "workDate": work_date,
                    "regularHours": float(day_info.get("regularHours", 0) or day_info.get("hours", 0)),
                    "otHours": float(day_info.get("ot", 0)),
                    "shiftCode": day_info.get("shiftTypeKey", "DAY"),
                    "dayType": day_info.get("dayType", "Ngày thường"),
                    "dayTypeKey": day_info.get("dayTypeKey", "weekday"),
                    "otMultiplier": float(day_info.get("otMultiplier", 1.0)),
                    "isHoliday": day_info.get("isHoliday", False),
                }
                timesheet_lines.append(line)

        return {
            "timesheets": timesheets,
            "timesheetLines": timesheet_lines,
            "metadata": {
                "project": self.project_name,
                "periodMonth": self.period_month,
                "periodYear": self.period_year,
                "totalEmployees": len(self.data),
                "exportedAt": datetime.now().isoformat()
            }
        }

    def get_db_payload(self) -> dict:
        """Get DB payload without creating file"""
        return self._prepare_db_payload()


def export_payroll_to_excel(parsed_data: list, project_name: str,
                           period_month: int, period_year: int,
                           output_path: str) -> dict:
    """Hàm tiện ích để export payroll"""
    manager = ExportManager(parsed_data, project_name, period_month, period_year)
    return manager.export_to_excel(output_path)
