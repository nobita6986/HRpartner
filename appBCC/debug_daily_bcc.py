# -*- coding: utf-8 -*-
"""Sum BCC daily breakdown per day to verify KE and KI"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.utils import get_column_letter

wb_bcc = openpyxl.load_workbook(r'C:\CodeApp\HrP\appBCC\docs\Actro\BCCActroT7.xlsx', data_only=True)
ws_bcc = wb_bcc.active

# Find A602010881
code = 'A602010881'
for row in range(3, ws_bcc.max_row + 1):
    c = ws_bcc.cell(row=row, column=2).value
    if c and str(c).strip() == code:
        print(f"A602010881 at row {row}\n")
        break

# Now scan daily cells - day 26 starts at col F (col 6 in 1-based, index 5)
# BCC Overtime structure: Day N starts at col idx = 5 + (N-26) * 9 (if N starts at 26)
# Layout per day: in, out, hours, OT1.5, 8h17, 20h22, đêm, 200% đêm, 210% CN
# 1-based col: in=N, out=N+1, hours=N+2, OT1.5=N+3, 8h17=N+4, 20h22=N+5, đêm=N+6, 200%đêm=N+7, 210%CN=N+8

print(f"{'Day':<5} | {'In':<7} | {'Out':<7} | {'Hours':<6} | {'OT1.5':<6} | {'Đêm':<6} | {'200%đêm':<8} | {'210%CN':<6}")
print("-" * 80)

days = [26, 27, 28, 29, 30, 31, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25]

total_hours = 0  # KE
total_ot_1_5 = 0  # KF
total_night = 0  # KI (includes 130% + 200%đêm)
total_night_200 = 0  # KH
total_sunday_210 = 0  # KJ

day_data_list = []
for day_num in days:
    # Calc day start column
    day_start = 6 + (day_num - 26) * 9 if day_num <= 25 else 6  # day 26 = col F=6
    
    in_time = ws_bcc.cell(row=row, column=day_start).value
    out_time = ws_bcc.cell(row=row, column=day_start+1).value
    hours = ws_bcc.cell(row=row, column=day_start+2).value
    ot_1_5 = ws_bcc.cell(row=row, column=day_start+3).value
    h_8_17 = ws_bcc.cell(row=row, column=day_start+4).value
    h_20_22 = ws_bcc.cell(row=row, column=day_start+5).value
    night = ws_bcc.cell(row=row, column=day_start+6).value
    night_200 = ws_bcc.cell(row=row, column=day_start+7).value
    cn_210 = ws_bcc.cell(row=row, column=day_start+8).value
    
    if any([in_time, out_time, ot_1_5, night, cn_210]):
        in_str = str(in_time).split(' ')[-1] if in_time else ''
        out_str = str(out_time).split(' ')[-1] if out_time else ''
        print(f"{day_num:<5} | {in_str:<7} | {out_str:<7} | {str(hours or 0):<6} | {str(ot_1_5 or 0):<6} | {str(night or 0):<6} | {str(night_200 or 0):<8} | {str(cn_210 or 0):<6}")
        
        day_data_list.append({
            'day': day_num,
            'in': in_time,
            'out': out_time,
            'hours': hours,
            'ot_1_5': ot_1_5,
            'night_130': night,
            'night_200': night_200,
            'cn_210': cn_210,
        })

# Sum all
print(f"\n=== BCC Daily Totals ===")
total_h_8_17_hours = 0  # normal hours (counts as 100%)
total_ot_1_5 = sum(d['ot_1_5'] or 0 for d in day_data_list)
total_night_130 = sum(d['night_130'] or 0 for d in day_data_list)
total_night_200 = sum(d['night_200'] or 0 for d in day_data_list)
total_cn_210 = sum(d['cn_210'] or 0 for d in day_data_list)
total_hours = sum(d['hours'] or 0 for d in day_data_list)

print(f"Total hours (col H): {total_hours}")
print(f"Total OT 1.5 (col I): {total_ot_1_5}")
print(f"Total Night 130% (col L): {total_night_130}")
print(f"Total Night 200% (col M): {total_night_200}")
print(f"Total CN 210% (col N): {total_cn_210}")

print(f"\n=== Comparing BCC Summary ===")
print(f"KC (sum OT-CN=210% + night-200%): {total_night_200 + total_cn_210}")
print(f"KD (số giờ ban ngày 130+150): should include hours + OT1.5 (roughly)")
print(f"KE: should match some total")
print(f"KF (OT 1.5× for SUMPRODUCT): {total_ot_1_5}")
print(f"KH (Night day for SUMPRODUCT): {total_night_200}")
print(f"KI (?): {total_night_130}")
print(f"KJ (Sunday for SUMPRODUCT): {total_cn_210}")

print(f"\n=== BCC vs HR ===")
print(f"  KC: BCC=0, HR=0 ✓")
print(f"  KD: BCC=88, HR=88 ✓")
print(f"  KE: BCC=20, HR=18 ✗ DIFF")
print(f"  KF: BCC=42, HR=42 ✓")
print(f"  KH: BCC=6, HR=6 ✓")
print(f"  KI: BCC=60, HR=54 ✗ DIFF")
print(f"  KJ: BCC=0, HR=0 ✓")

# Look at formulas to understand what KE/KI actually sum
wb_f = openpyxl.load_workbook(r'C:\CodeApp\HrP\appBCC\docs\Actro\LCNT7.xlsx', data_only=False)
ws_f = wb_f.active

print(f"\n=== HR Formulas (for understanding) ===")
formulas = {
    289: 'KC',
    290: 'KD',
    291: 'KE',
    292: 'KF',
    294: 'KH',
    295: 'KI',
    296: 'KJ',
    297: 'KK',
}
print("Row 10 headers:")
for col, name in formulas.items():
    h = ws_f.cell(row=10, column=col).value
    print(f"  {name} ({get_column_letter(col)}): {h}")

print("\nRow 13 multipliers:")
for col, name in formulas.items():
    v = ws_f.cell(row=13, column=col).value
    print(f"  {name} ({get_column_letter(col)}13): {v}")

# Row 11 (criteria)
print("\nRow 11 (criteria/categories):")
for col, name in formulas.items():
    v = ws_f.cell(row=11, column=col).value
    print(f"  {name} ({get_column_letter(col)}11): {v}")